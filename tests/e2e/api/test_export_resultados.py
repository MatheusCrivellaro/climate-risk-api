"""Testes e2e do export de resultados de estresse hídrico (Slice 20.1)."""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from climate_risk.core.ids import gerar_id
from climate_risk.domain.entidades.execucao import StatusExecucao
from climate_risk.infrastructure.db.conversores_tempo import datetime_para_iso
from climate_risk.infrastructure.db.modelos import (
    ExecucaoORM,
    MunicipioORM,
    ResultadoEstresseHidricoORM,
)
from climate_risk.interfaces.rotas import estresse_hidrico as rotas_estresse_hidrico


async def _inserir_municipio(
    sessionmaker: async_sessionmaker[AsyncSession],
    *,
    id_: int,
    nome: str,
    uf: str,
) -> None:
    agora_iso = datetime_para_iso(datetime.now(UTC))
    async with sessionmaker() as sessao:
        sessao.add(
            MunicipioORM(
                id=id_,
                nome=nome,
                nome_normalizado=nome.lower(),
                uf=uf,
                lat_centroide=None,
                lon_centroide=None,
                atualizado_em=agora_iso,
            )
        )
        await sessao.commit()


async def _inserir_execucao(
    sessionmaker: async_sessionmaker[AsyncSession],
    *,
    cenario: str = "rcp45",
) -> str:
    execucao_id = gerar_id("exec")
    agora_iso = datetime_para_iso(datetime.now(UTC))
    async with sessionmaker() as sessao:
        sessao.add(
            ExecucaoORM(
                id=execucao_id,
                cenario=cenario,
                variavel="pr+tas+evap",
                arquivo_origem="/tmp/pr.nc",
                tipo="estresse_hidrico",
                parametros="{}",
                status=StatusExecucao.COMPLETED,
                criado_em=agora_iso,
                concluido_em=agora_iso,
                job_id=None,
            )
        )
        await sessao.commit()
    return execucao_id


async def _inserir_resultado(
    sessionmaker: async_sessionmaker[AsyncSession],
    *,
    execucao_id: str,
    municipio_id: int,
    ano: int,
    cenario: str,
    frequencia: int = 10,
    intensidade: float = 12.5,
) -> None:
    rid = gerar_id("reh")
    agora_iso = datetime_para_iso(datetime.now(UTC))
    async with sessionmaker() as sessao:
        sessao.add(
            ResultadoEstresseHidricoORM(
                id=rid,
                execucao_id=execucao_id,
                municipio_id=municipio_id,
                ano=ano,
                cenario=cenario,
                frequencia_dias_secos_quentes=frequencia,
                intensidade_mm_dia=intensidade,
                criado_em=agora_iso,
            )
        )
        await sessao.commit()


@pytest.mark.asyncio
async def test_export_csv_status_e_content_type(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
) -> None:
    await _inserir_municipio(async_sessionmaker_, id_=3550308, nome="São Paulo", uf="SP")
    exec_id = await _inserir_execucao(async_sessionmaker_)
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp45",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "csv"}
    )
    assert resposta.status_code == 200
    assert resposta.headers["content-type"].startswith("text/csv")
    content_disposition = resposta.headers["content-disposition"]
    assert content_disposition.startswith("attachment;")
    assert ".csv" in content_disposition

    texto = resposta.content.decode("utf-8-sig")
    leitor = csv.reader(io.StringIO(texto))
    linhas = list(leitor)
    assert linhas[0][0] == "id"
    assert "nome_municipio" in linhas[0]
    assert any("São Paulo" in celula for celula in linhas[1])


@pytest.mark.asyncio
async def test_export_xlsx_abre_com_openpyxl(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
) -> None:
    await _inserir_municipio(async_sessionmaker_, id_=3550308, nome="São Paulo", uf="SP")
    exec_id = await _inserir_execucao(async_sessionmaker_)
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp45",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "xlsx"}
    )
    assert resposta.status_code == 200
    assert (
        resposta.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in resposta.headers["content-disposition"]
    assert ".xlsx" in resposta.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(resposta.content))
    ws = workbook.active
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "id"
    assert "nome_municipio" in headers
    valores_segunda_linha = [cell.value for cell in ws[2]]
    assert "São Paulo" in valores_segunda_linha


@pytest.mark.asyncio
async def test_export_json_lista_de_objetos(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
) -> None:
    await _inserir_municipio(async_sessionmaker_, id_=3550308, nome="São Paulo", uf="SP")
    exec_id = await _inserir_execucao(async_sessionmaker_)
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp45",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "json"}
    )
    assert resposta.status_code == 200
    assert resposta.headers["content-type"].startswith("application/json")
    assert "attachment;" in resposta.headers["content-disposition"]
    corpo = resposta.json()
    assert isinstance(corpo, list)
    assert len(corpo) == 1
    item = corpo[0]
    assert item["nome_municipio"] == "São Paulo"
    assert item["uf"] == "SP"
    assert item["execucao_id"] == exec_id


@pytest.mark.asyncio
async def test_export_filtro_execucao_id(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
) -> None:
    await _inserir_municipio(async_sessionmaker_, id_=3550308, nome="São Paulo", uf="SP")
    exec_a = await _inserir_execucao(async_sessionmaker_, cenario="rcp45")
    exec_b = await _inserir_execucao(async_sessionmaker_, cenario="rcp85")
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_a,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp45",
    )
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_b,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp85",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export",
        params={"formato": "json", "execucao_id": exec_b},
    )
    assert resposta.status_code == 200
    corpo = resposta.json()
    assert len(corpo) == 1
    assert corpo[0]["execucao_id"] == exec_b


@pytest.mark.asyncio
async def test_export_filtro_uf(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
) -> None:
    await _inserir_municipio(async_sessionmaker_, id_=3550308, nome="São Paulo", uf="SP")
    await _inserir_municipio(async_sessionmaker_, id_=3304557, nome="Rio de Janeiro", uf="RJ")
    exec_id = await _inserir_execucao(async_sessionmaker_)
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp45",
    )
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3304557,
        ano=2026,
        cenario="rcp45",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export",
        params={"formato": "json", "uf": "SP"},
    )
    assert resposta.status_code == 200
    corpo = resposta.json()
    assert len(corpo) == 1
    assert corpo[0]["uf"] == "SP"
    assert corpo[0]["nome_municipio"] == "São Paulo"


@pytest.mark.asyncio
async def test_export_sem_filtros_com_poucas_linhas_ok(
    cliente_api: AsyncClient,
) -> None:
    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "json"}
    )
    assert resposta.status_code == 200
    assert resposta.json() == []


@pytest.mark.asyncio
async def test_export_acima_do_limite_retorna_400(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Reduz o limite a 0 e insere uma linha — count > limite ⇒ 400."""
    monkeypatch.setattr(rotas_estresse_hidrico, "LIMITE_LINHAS_EXPORT", 0)
    await _inserir_municipio(async_sessionmaker_, id_=3550308, nome="São Paulo", uf="SP")
    exec_id = await _inserir_execucao(async_sessionmaker_)
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3550308,
        ano=2026,
        cenario="rcp45",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "csv"}
    )
    assert resposta.status_code == 400
    assert "linhas" in resposta.text.lower()


@pytest.mark.asyncio
async def test_export_caracteres_especiais_preservados_csv(
    cliente_api: AsyncClient,
    async_sessionmaker_: async_sessionmaker[AsyncSession],
) -> None:
    await _inserir_municipio(async_sessionmaker_, id_=3549904, nome="São José", uf="SC")
    exec_id = await _inserir_execucao(async_sessionmaker_)
    await _inserir_resultado(
        async_sessionmaker_,
        execucao_id=exec_id,
        municipio_id=3549904,
        ano=2026,
        cenario="rcp45",
    )

    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "csv"}
    )
    assert resposta.status_code == 200
    texto = resposta.content.decode("utf-8-sig")
    assert "São José" in texto


@pytest.mark.asyncio
async def test_export_formato_invalido_retorna_422(cliente_api: AsyncClient) -> None:
    resposta = await cliente_api.get(
        "/api/resultados/estresse-hidrico/export", params={"formato": "pdf"}
    )
    assert resposta.status_code == 422
